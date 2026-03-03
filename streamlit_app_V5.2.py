import streamlit as st
import pandas as pd
import numpy as np
import io
import math
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="MLS vs CAMA Data Comparison", page_icon="📊", layout="wide")

UNIQUE_ID_COLUMN = {'mls_col': 'Parcel Number', 'cama_col': 'PARID'}
COLUMNS_TO_COMPARE = [
    {'mls_col': 'Above Grade Finished Area', 'cama_col': 'SFLA'},
    {'mls_col': 'Bedrooms Total', 'cama_col': 'RMBED'},
    {'mls_col': 'Bathrooms Full', 'cama_col': 'FIXBATH'},
    {'mls_col': 'Bathrooms Half', 'cama_col': 'FIXHALF'},
]
COLUMNS_TO_COMPARE_SUM = [
    {'mls_col': 'Below Grade Finished Area', 'cama_cols': ['RECROMAREA', 'FINBSMTAREA', 'UFEATAREA']}
]
COLUMNS_TO_COMPARE_CATEGORICAL = [
    {'mls_col': 'Cooling', 'cama_col': 'HEAT', 'mls_check_contains': 'Central Air',
     'cama_expected_if_true': 1, 'cama_expected_if_false': 0, 'case_sensitive': False}
]
NUMERIC_TOLERANCE = 0.01
SKIP_ZERO_VALUES = True
ADDRESS_COLUMNS = {'address': 'Address', 'city': 'City', 'state': 'State or Province', 'zip': 'Postal Code'}
ZILLOW_URL_BASE = "https://www.zillow.com/homes/"
MLS_BATCH_SIZE = 35

MLS_TEXT_COLUMNS = {
    'Address', 'City', 'State or Province', 'State Or Province', 'Postal Code',
    'Public Remarks', 'Architectural Style', 'Lot Features', 'Buyer Financing',
    'Basement', 'Heating', 'Cooling', 'Additional Parcels Description',
    'Above Grade Finished Area Source', 'Below Grade Finished Area Source',
    'Lot Size Dimensions Source', 'Lot Size Units', 'Listing #',
    'MLS Status', 'Property Type', 'County', 'Subdivision'
}

def fill_mls_blanks(df):
    filled_count = 0
    for col in df.columns:
        if col in MLS_TEXT_COLUMNS:
            continue
        blanks = df[col].isna().sum()
        if blanks > 0:
            df[col] = df[col].fillna(0)
            filled_count += blanks
    return df, filled_count

def format_parcel_for_mls(parcel_id):
    try:
        return str(int(parcel_id)).zfill(8)
    except (ValueError, TypeError):
        return str(parcel_id).zfill(8)

def create_mls_batch_export(missing_parcels_df, batch_size=35):
    output = io.BytesIO()
    parcels = missing_parcels_df['Parcel_ID'].tolist()
    formatted = [format_parcel_for_mls(p) for p in parcels]
    batches = [formatted[i:i+batch_size] for i in range(0, len(formatted), batch_size)]
    total_batches = len(batches)
    summary_rows = []
    for bn, batch in enumerate(batches, 1):
        for p in batch:
            summary_rows.append({'Parcel_ID_Formatted': p, 'Batch': bn})
    summary_df = pd.DataFrame(summary_rows)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='All Parcels')
        for bn, batch in enumerate(batches, 1):
            pd.DataFrame({'Parcel_ID': batch}).to_excel(writer, index=False, sheet_name=f'Batch {bn}')
    output.seek(0)
    wb = load_workbook(output)
    batch_colors = ['DDEEFF','EEFFDD','FFEECC','FFE0E0','E8E0FF','E0F5FF','FFF0D0','F0FFE0','FFE8F5','E0E8FF']
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    ws_s = wb['All Parcels']
    for cell in ws_s[1]:
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = Alignment(horizontal='center')
    for row in ws_s.iter_rows(min_row=2, max_row=ws_s.max_row):
        bn = row[1].value
        if bn:
            fill = PatternFill(start_color=batch_colors[(bn-1)%len(batch_colors)], end_color=batch_colors[(bn-1)%len(batch_colors)], fill_type='solid')
            for cell in row:
                cell.fill = fill; cell.alignment = Alignment(horizontal='center')
    ws_s.column_dimensions['A'].width = 22; ws_s.column_dimensions['B'].width = 12
    for bn in range(1, total_batches+1):
        ws = wb[f'Batch {bn}']
        ws['A1'].fill = hdr_fill; ws['A1'].font = hdr_font; ws['A1'].alignment = Alignment(horizontal='center')
        ws['C1'] = f'Batch {bn} of {total_batches}  —  Copy column A and paste into MLS Parcel search'
        ws['C1'].font = Font(italic=True, color='555555')
        bd = [r[0].value for r in ws.iter_rows(min_row=2, max_row=ws.max_row) if r[0].value]
        ws['C2'] = f'{len(bd)} parcels in this batch'
        ws['C2'].font = Font(bold=True, color='1F4E79')
        color = batch_colors[(bn-1)%len(batch_colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value:
                    cell.fill = fill; cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(name='Courier New', size=11)
        ws.column_dimensions['A'].width = 16; ws.column_dimensions['C'].width = 65
    final_output = io.BytesIO()
    wb.save(final_output); final_output.seek(0)
    return final_output.getvalue()

def format_zillow_url(address, city, state, zip_code):
    if pd.isna(address) or pd.isna(city) or pd.isna(zip_code):
        return None
    import re
    ac = str(address).strip(); cc = str(city).strip(); zc = str(zip_code).strip().split('-')[0]
    ac = re.sub(r'\s+(Apt|Unit|#|Suite)\s*[\w-]*$','',ac,flags=re.IGNORECASE)
    af = re.sub(r'\s+','-',re.sub(r'[^\w\s-]','',ac))
    cf = re.sub(r'\s+','-',re.sub(r'[^\w\s-]','',cc))
    return f"{ZILLOW_URL_BASE}{af}-{cf}-OH-{zc}_rb/"

def values_equal(val1, val2):
    try:
        n1 = pd.to_numeric(val1, errors='raise'); n2 = pd.to_numeric(val2, errors='raise')
        if pd.isna(n1) and pd.isna(n2): return True
        elif pd.isna(n1) != pd.isna(n2): return False
        else: return np.isclose(n1, n2, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except (ValueError, TypeError):
        s1 = str(val1).strip().lower() if pd.notna(val1) else ''
        s2 = str(val2).strip().lower() if pd.notna(val2) else ''
        return s1 == s2

def categorical_match(mls_val, cama_val, mapping):
    ct = mapping.get('mls_check_contains',''); cs = mapping.get('case_sensitive',False)
    ms = str(mls_val).strip() if pd.notna(mls_val) else ''
    if not cs: ms = ms.lower(); ct = ct.lower()
    tf = ct in ms
    ec = mapping.get('cama_expected_if_true') if tf else mapping.get('cama_expected_if_false')
    try:
        cn = pd.to_numeric(cama_val, errors='coerce'); en = pd.to_numeric(ec, errors='coerce')
        if pd.isna(cn) and pd.isna(en): return True
        elif pd.isna(cn) or pd.isna(en): return False
        else: return np.isclose(cn, en, equal_nan=False, rtol=1e-9, atol=NUMERIC_TOLERANCE)
    except: return str(cama_val).strip().lower() == str(ec).strip().lower()

def calculate_difference(val1, val2):
    try:
        n1 = pd.to_numeric(val1, errors='raise'); n2 = pd.to_numeric(val2, errors='raise')
        if pd.isna(n1) or pd.isna(n2): return "N/A"
        return f"{n1-n2:,.2f}"
    except (ValueError, TypeError): return "Text difference"

def generate_mass_update_files(combined_df, user_initials, user_full_name):
    missing = [c for c in ['Parcel_ID','SALEKEY'] if c not in combined_df.columns]
    if missing: raise ValueError(f"Missing required columns: {missing}")
    lc = 'Listing_Number' if 'Listing_Number' in combined_df.columns else ('Listing #' if 'Listing #' in combined_df.columns else None)
    udf = combined_df.drop_duplicates(subset=['Parcel_ID'], keep='first').copy()
    sr = []; er = []
    for _, row in udf.iterrows():
        mp = row['Parcel_ID']; sk_str = str(row.get('SALEKEY','')).strip()
        ap_str = str(row.get('ADDITIONAL_PARCELS','')).strip()
        ln = 0
        if lc:
            ls = str(row.get(lc,'')).strip()
            if ls and ls != 'nan':
                try: ln = int(float(ls.rstrip(',').split(',')[0].strip()))
                except: ln = 0
        sks = [s.strip() for s in sk_str.rstrip(',').split(',') if s.strip()] if sk_str and sk_str != 'nan' else []
        aps = [p.strip() for p in ap_str.split(',') if p.strip()] if ap_str and ap_str != 'nan' else []
        for i, parcel in enumerate([mp] + aps):
            if i < len(sks):
                try:
                    si = int(sks[i]); pi = int(parcel)
                    sr.append({'PARID':pi,'SALEKEY':si,'USER11':ln,'SOURCE':0,'SALEVAL':0,'USER1':user_initials,'USER2':pd.Timestamp.now().strftime('%Y-%m-%d')})
                    er.append({'Change Type':'existing','appraiser':user_full_name,'parcelnum':pi,'comment':'','Review Status':'Reviewed','Determination':'','Est. Value Change':'','Last Changed Date/Time':pd.Timestamp.now().strftime('%m/%d/%Y %H:%M'),'Last Changed By':user_full_name})
                except: continue
    sdf = pd.DataFrame(sr)
    if not sdf.empty: sdf = sdf.drop_duplicates(subset=['SALEKEY'], keep='first').sort_values('SALEKEY').reset_index(drop=True)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w: sdf.to_excel(w, index=False, sheet_name='Sheet1')
    buf.seek(0)
    return buf.getvalue(), pd.DataFrame(er).to_csv(index=False)

def compare_data_enhanced(df_mls, df_cama, unique_id_col, cols_to_compare_mapping,
                          cols_to_compare_sum=None, cols_to_compare_categorical=None, window_id=None):
    if df_mls is None or df_cama is None:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    mls_id = unique_id_col.get('mls_col'); cama_id = unique_id_col.get('cama_col')
    if mls_id not in df_mls.columns: st.error(f"'{mls_id}' not in MLS"); return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if cama_id not in df_cama.columns: st.error(f"'{cama_id}' not in CAMA"); return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    dmr = df_mls.copy().rename(columns={mls_id: cama_id})
    matched_df = pd.merge(dmr, df_cama, on=cama_id, how='inner')
    merged_df = pd.merge(dmr, df_cama, on=cama_id, how='outer', indicator=True)
    pu = f"https://iasworld.starkcountyohio.gov/iasworld/Maintain/Transact.aspx?txtMaskedPin={{parcel_id}}&selYear=&userYear=&selJur=&chkShowHistory=False&chkShowChanges=&chkShowDeactivated=&PinValue={{parcel_id}}&pin=&trans_key=&windowId={window_id}&submitFlag=true&TransPopUp=&ACflag=False&ACflag2=False" if window_id else None
    mic=[]; mim=[]; vm=[]; pm=[]
    for _, row in merged_df.iterrows():
        rid = row.get(cama_id); ms = row.get('_merge')
        if ms == 'left_only':
            mic.append({'Parcel_ID':rid,'Listing_Number':row.get('Listing #',''),'Closed_Date':row.get('Closed Date','')})
        elif ms == 'right_only':
            mim.append({'Parcel_ID':rid})
        elif ms == 'both':
            ln=row.get('Listing #',''); sk=row.get('SALEKEY',''); np_=row.get('NOPAR',''); ap=row.get('ADDITIONAL_PARCELS','')
            addr=row.get('Address',''); city=row.get('City',''); state=row.get('State or Province',''); zc=row.get('Postal Code','')
            rmm=[]; fc=[]
            for m in cols_to_compare_mapping:
                mc=m['mls_col']; cc=m['cama_col']
                if mc not in merged_df.columns or cc not in merged_df.columns: continue
                mv=row.get(mc); cv=row.get(cc)
                if pd.isna(mv) or (isinstance(mv,str) and mv.strip()==''): continue
                if pd.isna(cv) or (isinstance(cv,str) and cv.strip()==''): continue
                fc.append(mc)
                if SKIP_ZERO_VALUES:
                    try:
                        mn=pd.to_numeric(mv,errors='coerce'); cn=pd.to_numeric(cv,errors='coerce')
                        if (pd.notna(mn) and mn==0) and (pd.notna(cn) and cn==0): continue
                    except: pass
                if not values_equal(mv, cv):
                    rmm.append({'Parcel_ID':rid,'NOPAR':np_,'ADDITIONAL_PARCELS':ap,'Listing_Number':ln,'SALEKEY':sk,'Address':addr,'City':city,'State':state,'Zip':zc,'Field_MLS':mc,'Field_CAMA':cc,'MLS_Value':mv,'CAMA_Value':cv,'Difference':calculate_difference(mv,cv),'Parcel_URL':pu.format(parcel_id=rid) if pu else '','Zillow_URL':format_zillow_url(addr,city,state,zc)})
            if cols_to_compare_sum:
                for m in cols_to_compare_sum:
                    mc=m['mls_col']; ccs=m['cama_cols']
                    if mc not in merged_df.columns or any(c not in merged_df.columns for c in ccs): continue
                    mv=row.get(mc)
                    if pd.isna(mv) or (isinstance(mv,str) and mv.strip()==''): continue
                    csum=0; acb=True
                    for c in ccs:
                        v=row.get(c)
                        if pd.notna(v): acb=False; csum+=pd.to_numeric(v,errors='coerce')
                    if acb: continue
                    fc.append(mc)
                    if SKIP_ZERO_VALUES:
                        try:
                            mn=pd.to_numeric(mv,errors='coerce')
                            if (pd.notna(mn) and mn==0) and csum==0: continue
                        except: pass
                    if not values_equal(mv, csum):
                        rmm.append({'Parcel_ID':rid,'NOPAR':np_,'ADDITIONAL_PARCELS':ap,'Listing_Number':ln,'SALEKEY':sk,'Address':addr,'City':city,'State':state,'Zip':zc,'Field_MLS':mc,'Field_CAMA':f"SUM({', '.join(ccs)})",'MLS_Value':mv,'CAMA_Value':csum,'Difference':calculate_difference(mv,csum),'Parcel_URL':pu.format(parcel_id=rid) if pu else '','Zillow_URL':format_zillow_url(addr,city,state,zc)})
            if cols_to_compare_categorical:
                for m in cols_to_compare_categorical:
                    mc=m['mls_col']; cc=m['cama_col']
                    if mc not in merged_df.columns or cc not in merged_df.columns: continue
                    mv=row.get(mc); cv=row.get(cc)
                    if pd.isna(mv) or (isinstance(mv,str) and mv.strip()==''): continue
                    if pd.isna(cv) or (isinstance(cv,str) and cv.strip()==''): continue
                    fc.append(mc); im=categorical_match(mv,cv,m)
                    ct=m.get('mls_check_contains',''); cs=m.get('case_sensitive',False)
                    ms2=str(mv).strip().lower() if pd.notna(mv) else ''
                    ct2=ct.lower() if not cs else ct
                    ec=m.get('cama_expected_if_true') if ct2 in ms2 else m.get('cama_expected_if_false')
                    if not im:
                        rmm.append({'Parcel_ID':rid,'NOPAR':np_,'ADDITIONAL_PARCELS':ap,'Listing_Number':ln,'SALEKEY':sk,'Address':addr,'City':city,'State':state,'Zip':zc,'Field_MLS':mc,'Field_CAMA':cc,'MLS_Value':mv,'CAMA_Value':cv,'Expected_CAMA_Value':ec,'Match_Rule':f"If '{ct}' in {mc}, then {cc} should be {m.get('cama_expected_if_true')}, else {m.get('cama_expected_if_false')}",'Parcel_URL':pu.format(parcel_id=rid) if pu else '','Zillow_URL':format_zillow_url(addr,city,state,zc)})
            if not rmm and fc:
                pm.append({'Parcel_ID':rid,'NOPAR':np_,'ADDITIONAL_PARCELS':ap,'Listing_Number':ln,'SALEKEY':sk,'Address':addr,'City':city,'State':state,'Zip':zc,'Fields_Compared':len(fc),'Fields_List':', '.join(fc),'Parcel_URL':pu.format(parcel_id=rid) if pu else '','Zillow_URL':format_zillow_url(addr,city,state,zc)})
            vm.extend(rmm)
    return pd.DataFrame(mic), pd.DataFrame(mim), pd.DataFrame(vm), matched_df, pd.DataFrame(pm)

def create_excel_with_hyperlinks(df, sheet_name='Sheet1'):
    out=io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as w: df.to_excel(w, index=False, sheet_name=sheet_name)
    out.seek(0); wb=load_workbook(out); ws=wb[sheet_name]
    cols=list(df.columns)
    if 'Parcel_ID' in cols and 'Parcel_URL' in cols:
        pi=cols.index('Parcel_ID')+1; ui=cols.index('Parcel_URL')+1
        for r in range(2,len(df)+2):
            cell=ws.cell(row=r,column=pi); url=ws.cell(row=r,column=ui).value
            if url and str(url).strip() and str(url)!='nan': cell.hyperlink=url; cell.style='Hyperlink'
    if 'Address' in cols and 'Zillow_URL' in cols:
        ai=cols.index('Address')+1; zi=cols.index('Zillow_URL')+1
        for r in range(2,len(df)+2):
            cell=ws.cell(row=r,column=ai); url=ws.cell(row=r,column=zi).value
            if url and str(url).strip() and str(url)!='nan': cell.hyperlink=url; cell.style='Hyperlink'
    if 'Parcel_URL' in cols: ws.delete_cols(cols.index('Parcel_URL')+1)
    rc=[c for c in cols if c!='Parcel_URL']
    if 'Zillow_URL' in rc: ws.delete_cols(rc.index('Zillow_URL')+1)
    fo=io.BytesIO(); wb.save(fo); fo.seek(0)
    return fo.getvalue()

def create_zip_with_all_reports(dmc, dmm, dvm, dpm, city_df=None):
    import zipfile
    ts=datetime.now().strftime("%Y-%m-%d"); zb=io.BytesIO()
    with zipfile.ZipFile(zb,'w',zipfile.ZIP_DEFLATED) as zf:
        if not dmc.empty: zf.writestr(f"missing_in_CAMA_{ts}.xlsx", create_excel_with_hyperlinks(dmc,'Missing in CAMA'))
        if not dmm.empty: zf.writestr(f"missing_in_MLS_{ts}.xlsx", create_excel_with_hyperlinks(dmm,'Missing in MLS'))
        if not dvm.empty: zf.writestr(f"value_mismatches_{ts}.xlsx", create_excel_with_hyperlinks(dvm,'Value Mismatches'))
        if not dpm.empty: zf.writestr(f"perfect_matches_{ts}.xlsx", create_excel_with_hyperlinks(dpm,'Perfect Matches'))
        if city_df is not None and not city_df.empty: zf.writestr(f"city_match_statistics_{ts}.csv", city_df.to_csv(index=False))
    zb.seek(0); return zb.getvalue()

# ══════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════

st.title("📊 MLS vs CAMA Data Comparison Tool")
st.markdown("Compare MLS and CAMA property data to identify discrepancies and perfect matches.")

with st.sidebar:
    st.header("⚙️ Configuration")
    st.subheader("WindowId Setup")
    st.markdown("1. Log into iasWorld\n2. Search any parcel\n3. Copy `windowId` from the URL")
    window_id = st.text_input("Enter WindowId", value="638981240146803746")
    st.divider()
    st.subheader("Comparison Settings")
    tolerance = st.number_input("Numeric Tolerance", value=0.01, format="%.4f")
    skip_zeros = st.checkbox("Skip Zero Values", value=True, help="Skips comparison when BOTH values are 0")
    st.divider()
    st.subheader("MLS Batch Size")
    batch_size = st.number_input("Parcels per Batch", value=35, min_value=1, max_value=100)

NUMERIC_TOLERANCE = tolerance
SKIP_ZERO_VALUES = skip_zeros
MLS_BATCH_SIZE = int(batch_size)

st.header("📁 Upload Data Files")
col1, col2 = st.columns(2)
with col1:
    mls_file = st.file_uploader("Upload MLS Data (Excel)", type=['xlsx','xls'],
                                 help="Blanks auto-filled with 0 on upload — no CTRL+H needed")
with col2:
    cama_file = st.file_uploader("Upload CAMA Data (Excel)", type=['xlsx','xls'])

if mls_file and cama_file:
    with st.spinner("Loading data files..."):
        try:
            df_mls_raw = pd.read_excel(mls_file)
            df_cama = pd.read_excel(cama_file)
            df_mls, filled_count = fill_mls_blanks(df_mls_raw)
            if filled_count > 0:
                st.success(f"✅ MLS loaded — {filled_count:,} blank cells auto-filled with 0 (no CTRL+H needed)")
            else:
                st.success("✅ Files loaded successfully!")
        except Exception as e:
            st.error(f"Error loading files: {e}"); st.stop()

    st.header("📊 Data Summary")
    c1,c2,c3 = st.columns(3)
    c1.metric("MLS Records", len(df_mls)); c2.metric("CAMA Records", len(df_cama)); c3.metric("Tolerance", NUMERIC_TOLERANCE)

    c1,c2 = st.columns([3,1])
    with c1: run_button = st.button("🔍 Run Comparison", type="primary", use_container_width=True)
    with c2:
        if st.session_state.get('comparison_complete', False):
            if st.button("🔄 Clear Results", use_container_width=True):
                for k in ['df_missing_cama','df_missing_mls','df_value_mismatches','matched_df','df_perfect_matches','comparison_complete','city_comparison']:
                    st.session_state.pop(k, None)
                st.rerun()

    if run_button:
        with st.spinner("Comparing data..."):
            dmc,dmm,dvm,mdf,dpm = compare_data_enhanced(df_mls, df_cama, UNIQUE_ID_COLUMN, COLUMNS_TO_COMPARE,
                                                         cols_to_compare_sum=COLUMNS_TO_COMPARE_SUM,
                                                         cols_to_compare_categorical=COLUMNS_TO_COMPARE_CATEGORICAL,
                                                         window_id=window_id)
            st.session_state.update({'df_missing_cama':dmc,'df_missing_mls':dmm,'df_value_mismatches':dvm,
                                     'matched_df':mdf,'df_perfect_matches':dpm,'comparison_complete':True})

    if st.session_state.get('comparison_complete', False):
        dmc=st.session_state['df_missing_cama']; dmm=st.session_state['df_missing_mls']
        dvm=st.session_state['df_value_mismatches']; mdf=st.session_state['matched_df']
        dpm=st.session_state['df_perfect_matches']

        st.success("✅ Comparison complete.")

        st.header("📈 Results Summary")
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("✅ Matched", len(mdf)); c2.metric("❌ Missing in CAMA", len(dmc))
        c3.metric("❌ Missing in MLS", len(dmm)); c4.metric("⚠️ Value Mismatches", len(dvm))
        c1,c2 = st.columns(2)
        c1.metric("✅ Perfect Matches", len(dpm))
        if not dvm.empty: c2.metric("📊 Fields with Mismatches", dvm['Field_MLS'].nunique())

        # ── MISSING IN MLS BATCH EXPORT ──
        if not dmm.empty:
            st.header("🔍 Missing in MLS — MLS Search Batches")
            nb = math.ceil(len(dmm) / MLS_BATCH_SIZE)
            st.markdown(
                f"**{len(dmm)} parcels** not found in MLS. Download below — parcels are **zero-padded to 8 digits** "
                f"and split into **{nb} batch{'es' if nb>1 else ''}** of up to {MLS_BATCH_SIZE}, "
                f"ready to copy-paste into the MLS parcel search."
            )
            c1,c2,c3 = st.columns(3)
            c1.metric("Missing Parcels", len(dmm)); c2.metric("Batches Needed", nb); c3.metric("Per Batch", MLS_BATCH_SIZE)
            ts=datetime.now().strftime("%m%d%y")
            st.download_button(
                label=f"📥 Download MLS Search Batches ({nb} batches · {len(dmm)} parcels)",
                data=create_mls_batch_export(dmm, MLS_BATCH_SIZE),
                file_name=f"MLS_Search_Batches_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            with st.expander("👁️ Preview formatted parcels"):
                prev=dmm.copy()
                prev['Parcel_ID_Formatted']=prev['Parcel_ID'].apply(format_parcel_for_mls)
                prev['Batch']=[(i//MLS_BATCH_SIZE)+1 for i in range(len(prev))]
                st.dataframe(prev[['Parcel_ID','Parcel_ID_Formatted','Batch']], use_container_width=True, hide_index=True)

        # City stats
        st.header("📊 CAMA Parcel Match Statistics")
        tcp=len(df_cama); mp=len(mdf); mr=(mp/tcp*100) if tcp>0 else 0
        c1,c2,c3=st.columns(3)
        c1.metric("Total CAMA Parcels",f"{tcp:,}"); c2.metric("Found in MLS",f"{mp:,}"); c3.metric("Match Rate",f"{mr:.2f}%")

        st.subheader("Match Rate by City")
        cc_col = 'CITYNAME' if 'CITYNAME' in df_cama.columns else ('City' if 'City' in df_cama.columns else None)
        cama_id_col = UNIQUE_ID_COLUMN.get('cama_col')
        if cc_col and not mdf.empty and cc_col in mdf.columns:
            cc_df=df_cama.groupby(cc_col)[cama_id_col].count().reset_index(); cc_df.columns=['City','Total_CAMA_Parcels']
            mc_df=mdf.groupby(cc_col)[cama_id_col].count().reset_index(); mc_df.columns=['City','Matched_Parcels']
            cyc=pd.merge(cc_df,mc_df,on='City',how='left')
            cyc['Matched_Parcels']=cyc['Matched_Parcels'].fillna(0).astype(int)
            cyc['Match_Rate']=(cyc['Matched_Parcels']/cyc['Total_CAMA_Parcels']*100).round(2)
            cyc['Not_Matched']=cyc['Total_CAMA_Parcels']-cyc['Matched_Parcels']
            cyc=cyc.sort_values('Total_CAMA_Parcels',ascending=False)
            st.session_state['city_comparison']=cyc
            st.dataframe(cyc[['City','Total_CAMA_Parcels','Matched_Parcels','Not_Matched','Match_Rate']], use_container_width=True, hide_index=True)
            st.download_button("📥 Download City Statistics (CSV)", data=cyc.to_csv(index=False),
                               file_name=f"city_match_statistics_{datetime.now().strftime('%Y-%m-%d')}.csv", mime="text/csv")
            c1,c2=st.columns(2)
            with c1:
                st.markdown("**Top 10 Cities by CAMA Parcels**"); tc=cyc.head(10)
                st.bar_chart(tc.set_index('City')[['Matched_Parcels','Not_Matched']])
            with c2:
                st.markdown("**Match Rate by City (Top 10)**")
                st.bar_chart(tc[['City','Match_Rate']].set_index('City'))
        else:
            st.info("ℹ️ City info not available")

        if not dvm.empty:
            st.subheader("📊 Mismatches by Field")
            st.bar_chart(dvm['Field_MLS'].value_counts())

        st.header("📋 Data Preview")
        t1,t2,t3,t4=st.tabs(["Missing in CAMA","Missing in MLS","Value Mismatches","Perfect Matches"])
        with t1:
            if not dmc.empty:
                st.dataframe(dmc, use_container_width=True)
            else:
                st.info("No records missing in CAMA")
        with t2:
            if not dmm.empty:
                st.dataframe(dmm, use_container_width=True)
            else:
                st.info("No records missing in MLS")
        with t3:
            if not dvm.empty:
                st.dataframe(dvm, use_container_width=True)
            else:
                st.info("No value mismatches")
        with t4:
            if not dpm.empty:
                st.dataframe(dpm, use_container_width=True)
            else:
                st.info("No perfect matches")

        st.header("📥 Download Reports")
        st.markdown("### 📦 All Reports")
        ts=datetime.now().strftime("%Y-%m-%d"); cyc2=st.session_state.get('city_comparison',None)
        st.download_button("📦 Download All Reports (ZIP)",
                           data=create_zip_with_all_reports(dmc,dmm,dvm,dpm,cyc2),
                           file_name=f"MLS_CAMA_Comparison_All_Reports_{ts}.zip",
                           mime="application/zip", use_container_width=True)

        st.markdown("### 📄 Individual Reports")
        c1,c2=st.columns(2)
        with c1:
            if not dmc.empty: st.download_button("📄 Missing in CAMA", data=create_excel_with_hyperlinks(dmc,'Missing in CAMA'), file_name=f"missing_in_CAMA_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if not dvm.empty: st.download_button("⚠️ Value Mismatches", data=create_excel_with_hyperlinks(dvm,'Value Mismatches'), file_name=f"value_mismatches_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c2:
            if not dmm.empty: st.download_button("📄 Missing in MLS", data=create_excel_with_hyperlinks(dmm,'Missing in MLS'), file_name=f"missing_in_MLS_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if not dpm.empty: st.download_button("✅ Perfect Matches", data=create_excel_with_hyperlinks(dpm,'Perfect Matches'), file_name=f"perfect_matches_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.header("🔄 CAMA Mass Updates")
        if not dpm.empty or not dvm.empty:
            with st.expander("⚙️ Generate CAMA Mass Update Files", expanded=False):
                st.markdown("Generates:\n1. **SALETAB_MassUpdate** — SALE tab\n2. **MassEntrance** — ENTRANCE tab")
                c1,c2=st.columns(2)
                with c1: ui=st.text_input("Your Initials (e.g., JMJ)", value="JMJ")
                with c2: ufn=st.text_input("Your Full Name (e.g., Jason Jeffries)", value="Jason Jeffries")
                if st.button("🔄 Generate Mass Update Files", type="primary"):
                    cdf=pd.concat([dpm,dvm],ignore_index=True)
                    if cdf.empty: st.error("No data available")
                    else:
                        st.info(f"📋 {len(cdf)} combined records")
                        if 'Listing_Number' not in cdf.columns:
                            st.error("❌ Listing_Number not found!"); st.write("Columns:", cdf.columns.tolist())
                        else:
                            nn=cdf['Listing_Number'].notna().sum()
                            st.info(f"✔ Listing_Number: {nn}/{len(cdf)} non-null")
                            with st.expander("📊 Sample"): st.dataframe(cdf[['Parcel_ID','Listing_Number']].head(5))
                        sd,ed=generate_mass_update_files(cdf,ui,ufn)
                        st.success(f"✅ Files generated for {len(cdf)} records")
                        c1,c2=st.columns(2); fts=datetime.now().strftime("%m%d%y")
                        with c1: st.download_button("📥 SALETAB_MassUpdate", data=sd, file_name=f"SALETAB_MassUpdate_{fts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        with c2: st.download_button("📥 MassEntrance", data=ed, file_name=f"MassEntrance{fts}.csv", mime="text/csv")
        else:
            st.info("ℹ️ Run comparison first.")

else:
    st.info("👆 Upload both MLS and CAMA files to begin.")
    with st.expander("ℹ️ Expected Data Format"):
        st.markdown("""
        **MLS Data:** `Parcel Number` · `Above Grade Finished Area` · `Bedrooms Total` · `Bathrooms Full` · `Bathrooms Half` · `Below Grade Finished Area` · `Cooling` · `Address` · `City` · `State or Province` · `Postal Code`

        > Blank cells are **auto-filled with 0** on upload — no CTRL+H needed.

        **CAMA Data:** `PARID` · `NOPAR` · `CITYNAME` · `SFLA` · `RMBED` · `FIXBATH` · `FIXHALF` · `RECROMAREA` · `FINBSMTAREA` · `UFEATAREA` · `HEAT` · `SALEKEY`
        """)

st.divider()
st.caption("MLS vs CAMA Comparison Tool | Built with Streamlit")
